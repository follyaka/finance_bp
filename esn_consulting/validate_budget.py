"""
Script de validation du modèle Budget 2026
Vérifie l'intégrité du modèle selon les standards Transaction Services
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys
import glob
import os


def validate_budget_model(filename):
    """Validate the budget model integrity"""

    print(f"\n{'='*80}")
    print(f"VALIDATION DU MODÈLE: {filename}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(filename)

    validation_results = []
    errors = []
    warnings = []

    # ========================
    # CHECK 1: Worksheets exist
    # ========================
    print("✓ CHECK 1: Structure des onglets")
    required_sheets = ['Hypothèses', "Chiffre d'affaires"]
    for sheet_name in required_sheets:
        if sheet_name in wb.sheetnames:
            validation_results.append(f"  ✓ Onglet '{sheet_name}' présent")
        else:
            errors.append(f"  ✗ ERREUR: Onglet '{sheet_name}' manquant")

    # ========================
    # CHECK 2: Hypothèses values are reasonable
    # ========================
    print("\n✓ CHECK 2: Validation des hypothèses")
    ws_hypo = wb['Hypothèses']

    # Check commercial team
    nb_commerciaux = ws_hypo['B8'].value
    if nb_commerciaux and nb_commerciaux > 0:
        validation_results.append(f"  ✓ Nombre de commerciaux: {nb_commerciaux}")
    else:
        errors.append("  ✗ ERREUR: Nombre de commerciaux invalide")

    # Check ramp-up
    rampup_values = [ws_hypo[f'B{i}'].value for i in [9, 10, 11, 12]]
    if all(v is not None and v > 0 for v in rampup_values):
        validation_results.append(f"  ✓ Ramp-up configuré: {rampup_values}")
    else:
        errors.append("  ✗ ERREUR: Valeurs de ramp-up invalides")

    # Check transformation rate
    taux_transfo = ws_hypo['B13'].value
    if taux_transfo and 0 < taux_transfo <= 1:
        validation_results.append(f"  ✓ Taux de transformation: {taux_transfo:.1%}")
    else:
        warnings.append(f"  ⚠ Taux de transformation hors norme: {taux_transfo}")

    # Check consultants
    nb_consultants = ws_hypo['B16'].value
    if nb_consultants and nb_consultants > 0:
        validation_results.append(f"  ✓ Nombre de consultants: {nb_consultants}")
    else:
        errors.append("  ✗ ERREUR: Nombre de consultants invalide")

    # Check TJM
    tjm = ws_hypo['B17'].value
    if tjm and tjm > 0:
        validation_results.append(f"  ✓ TJM: {tjm:,.0f} €")
    else:
        errors.append("  ✗ ERREUR: TJM invalide")

    # Check TACE
    tace = ws_hypo['B18'].value
    if tace and 0 < tace <= 1:
        validation_results.append(f"  ✓ TACE: {tace:.1%}")
    else:
        warnings.append(f"  ⚠ TACE hors norme: {tace}")

    # Check mission duration
    duree_mission = ws_hypo['B19'].value
    if duree_mission and duree_mission > 0:
        validation_results.append(f"  ✓ Durée mission: {duree_mission} jours")
    else:
        errors.append("  ✗ ERREUR: Durée mission invalide")

    # ========================
    # CHECK 3: Formulas in CA worksheet
    # ========================
    print("\n✓ CHECK 3: Validation des formules")
    ws_ca = wb["Chiffre d'affaires"]

    # Check that key cells contain formulas, not hardcoded values
    formula_checks = [
        ('B7', 'Ramp-up Commercial 1 - Janvier'),
        ('B13', 'BC générés - Janvier'),
        ('B15', 'Factures signées - Janvier'),
        ('B17', 'CA facturé - Janvier'),
        ('B21', 'Jours ouvrés - Janvier'),
        ('B24', 'CA production MAX - Janvier'),
        ('B28', 'CA TOTAL MENSUEL - Janvier'),
    ]

    formulas_ok = 0
    for cell_ref, description in formula_checks:
        cell = ws_ca[cell_ref]
        if hasattr(cell, 'value') and isinstance(getattr(cell, 'value', None), str) and cell.value.startswith('='):
            formulas_ok += 1
        elif cell.data_type == 'f':  # Formula type
            formulas_ok += 1
        else:
            # Try to check if it's a formula cell
            if cell.value is not None:
                warnings.append(f"  ⚠ {cell_ref} ({description}): Vérifier si c'est bien une formule")

    if formulas_ok >= 5:
        validation_results.append(f"  ✓ Formules détectées dans les cellules clés ({formulas_ok} sur {len(formula_checks)})")
    else:
        warnings.append(f"  ⚠ Peu de formules détectées ({formulas_ok} sur {len(formula_checks)})")

    # ========================
    # CHECK 4: CA calculations
    # ========================
    print("\n✓ CHECK 4: Cohérence des calculs CA")

    # Get CA values for each month
    ca_values = []
    for col in range(2, 14):  # B to M (months)
        col_letter = get_column_letter(col)
        ca_cell = ws_ca[f'{col_letter}28']  # CA TOTAL MENSUEL row
        if ca_cell.value is not None:
            try:
                ca_values.append(float(ca_cell.value))
            except (ValueError, TypeError):
                ca_values.append(0)

    if len(ca_values) == 12:
        validation_results.append(f"  ✓ 12 mois de CA calculés")

        total_ca = sum(ca_values)
        validation_results.append(f"  ✓ CA total annuel: {total_ca:,.0f} €")

        # Check for negative values
        if all(v >= 0 for v in ca_values):
            validation_results.append("  ✓ Aucune valeur négative détectée")
        else:
            errors.append("  ✗ ERREUR: Valeurs négatives détectées dans le CA")

        # Check for progression
        if ca_values[-1] > ca_values[0]:
            validation_results.append(f"  ✓ Progression CA: {ca_values[0]:,.0f} € → {ca_values[-1]:,.0f} €")
        else:
            warnings.append("  ⚠ Pas de progression du CA sur l'année")

    else:
        errors.append(f"  ✗ ERREUR: Nombre de mois incorrect ({len(ca_values)} au lieu de 12)")

    # ========================
    # CHECK 5: Production capacity
    # ========================
    print("\n✓ CHECK 5: Vérification capacité production")

    ca_prod_max_values = []
    for col in range(2, 14):
        col_letter = get_column_letter(col)
        prod_cell = ws_ca[f'{col_letter}24']  # CA production MAX row
        if prod_cell.value is not None:
            try:
                ca_prod_max_values.append(float(prod_cell.value))
            except (ValueError, TypeError):
                ca_prod_max_values.append(0)

    if len(ca_prod_max_values) == 12 and len(ca_values) == 12:
        capacity_check_ok = True
        for i, (ca, ca_max) in enumerate(zip(ca_values, ca_prod_max_values)):
            if ca > ca_max * 1.01:  # Allow 1% tolerance
                capacity_check_ok = False
                errors.append(f"  ✗ ERREUR Mois {i+1}: CA ({ca:,.0f}€) > Capacité prod ({ca_max:,.0f}€)")

        if capacity_check_ok:
            validation_results.append("  ✓ CA ≤ Capacité production pour tous les mois")

        avg_utilization = sum(ca_values) / sum(ca_prod_max_values) if sum(ca_prod_max_values) > 0 else 0
        validation_results.append(f"  ✓ Taux d'utilisation moyen: {avg_utilization:.1%}")

    # ========================
    # CHECK 6: Formatting
    # ========================
    print("\n✓ CHECK 6: Validation du formatage")

    # Check header cells have proper formatting
    header_cell = ws_ca['A1']
    if header_cell.fill.start_color.rgb:
        validation_results.append("  ✓ En-têtes formatés")

    # ========================
    # SUMMARY
    # ========================
    print(f"\n{'='*80}")
    print("RÉSUMÉ DE LA VALIDATION")
    print(f"{'='*80}\n")

    for result in validation_results:
        print(result)

    if warnings:
        print(f"\n⚠ AVERTISSEMENTS ({len(warnings)}):")
        for warning in warnings:
            print(warning)

    if errors:
        print(f"\n✗ ERREURS ({len(errors)}):")
        for error in errors:
            print(error)
        print(f"\n{'='*80}")
        print("STATUT: ✗ ÉCHEC - Corrections requises")
        print(f"{'='*80}\n")
        return False
    else:
        print(f"\n{'='*80}")
        print("STATUT: ✓ SUCCÈS - Modèle validé")
        print(f"{'='*80}\n")
        return True


if __name__ == "__main__":
    # Find the most recent budget file
    budget_files = glob.glob("Budget_CA_2026_*.xlsx")

    if not budget_files:
        print("✗ ERREUR: Aucun fichier Budget_CA_2026_*.xlsx trouvé")
        sys.exit(1)

    # Sort by modification time (most recent first)
    budget_files.sort(key=os.path.getmtime, reverse=True)
    latest_file = budget_files[0]

    success = validate_budget_model(latest_file)
    sys.exit(0 if success else 1)
